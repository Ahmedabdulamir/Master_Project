import pandas as pd
import os
import x_section_w
import x_section
import y_intergration_w
import xlsxwriter
import openpyxl as xl
import shutil




def _opExcel(data, data_type, Path, Direction, Moment):
    
    if not os.path.exists('xlsx/' + Path + '/'+ Path + Direction + '.xlsm'):
        if not os.path.exists('xlsx/' + Path):
            os.mkdir('xlsx/' + Path)
                    
        base ='xlsx/Sheet.xlsm'
        copy ='xlsx/' + Path + '/'+ Path + Direction + '.xlsm'
        shutil.copy(base,copy)

        wb = xl.load_workbook('xlsx/' + Path + '/'+ Path + Direction + '.xlsm', read_only=False, keep_vba=True)
        ws = wb['Default']
    else:
        wb = xl.load_workbook('xlsx/' + Path + '/'+ Path + Direction + '.xlsm', read_only=False, keep_vba=True)
        ws = wb['Default']
    
    data_types = ['M', 'M_avg', 'M_h','V', 'Y']
    row_index = ['A10001', 'A10002', 'A10003','A10004','A10005']
    col_index = ['B10001', 'B10002', 'B10003','B10004','B10005']
    index = data_types.index(data_type)

    col = ws[col_index[index]].value
    row = ws[row_index[index]].value

    print(data)
    print(col)
    


    for i in range(row, row + len(data)):
        ws.cell(row=i,column=col).value = data[i-row]
        if index == 1 and i > 2:
            coor1 = _cell(ws.cell(row=i,column=col))
            coor2 = _cell(ws.cell(row=8,column=col))
           
            ws.cell(row=i+8,column=col).value = '=ROUND('+ coor2 + '/' + coor1 + ',2)'
            ws.cell(row=int(col/10)+6,column=i+272).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/2'
            ws.cell(row=int(col/10)+100,column=i+272).value = '='+ _cell(ws.cell(row=i+8 ,column=col)) + '/-2'
            ws.cell(row=int(col/10)+6,column=274).value = '='+ _cell(ws.cell(row=10 ,column=col))  
            ws.cell(row=int(col/10)+100,column=274).value = '='+ _cell(ws.cell(row=10 ,column=col))  

    ws[col_index[index]].value = int(col)+2
    if index == 1 or index == 2:
       ws[col_index[index]].value = int(col)+10 
    ws['A10006'] = str(Path)
        
    wb.save('xlsx/' + Path + '/'+ Path + Direction + '.xlsm')


def _cell(coor):
    coor = str(coor)
    coor = coor.rstrip('>')
    coor = coor.split(".")
    return str(coor[1])